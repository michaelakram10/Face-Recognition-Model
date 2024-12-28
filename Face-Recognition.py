import cv2
import numpy as np
from keras_facenet import FaceNet  # Pre-trained FaceNet model for embeddings
from mtcnn import MTCNN  # MTCNN for face detection
import pickle  # For saving/loading the database
from deepface import DeepFace  # For emotion detection
import tkinter as tk
from tkinter import filedialog  # For file dialog
from openpyxl import Workbook, load_workbook  # For working with Excel sheets
from datetime import datetime  # For timestamping attendance

# Initialize FaceNet and MTCNN
embedder = FaceNet()
detector = MTCNN()

# File path for saving/loading the database and attendance log
DATABASE_FILE = "known_faces.pkl"
ATTENDANCE_FILE = "attendance_log.xlsx"

# Function to load the database
def load_database():
    """Load the known faces database from a file."""
    try:
        with open(DATABASE_FILE, "rb") as file:
            return pickle.load(file)
    except FileNotFoundError:
        print("Database file not found. Starting with an empty database.")
        return {}

# Function to save the database
def save_database():
    """Save the known faces database to a file."""
    with open(DATABASE_FILE, "wb") as file:
        pickle.dump(known_faces, file)
    print("Database saved successfully.")

# Load the database at the start
known_faces = load_database()

# Initialize or load the attendance Excel sheet
def initialize_attendance():
    """Initialize the attendance Excel sheet if it doesn't exist."""
    try:
        workbook = load_workbook(ATTENDANCE_FILE)
        print("Attendance file found.")
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Attendance"
        sheet.append(["Name", "Timestamp", "Action"])
        workbook.save(ATTENDANCE_FILE)
        print("Attendance file created.")
    return workbook

attendance_workbook = initialize_attendance()

# Function to check if a face already exists in the database
def is_face_in_database(embedding):
    """Check if a face embedding already exists in the database."""
    for name, db_data in known_faces.items():
        db_embedding, _ = db_data  # Unpack the embedding and emotion data
        distance = np.linalg.norm(embedding - db_embedding)
        if distance < 1.0:  # Similarity threshold
            return name, distance
    return None, None

# Function to add faces to the database
def add_to_database_from_file(file_path):
    """Add faces from an image file to the database."""
    image = cv2.imread(file_path)
    if image is None:
        print("Error loading image. Please try again.")
        return

    rgb_image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    detections = detector.detect_faces(rgb_image)

    if len(detections) == 0:
        print("No face detected in the image. Please try again.")
        return

    for detection in detections:
        x, y, w, h = detection['box']
        x, y = max(0, x), max(0, y)  # Ensure coordinates are non-negative
        cropped_face = rgb_image[y:y+h, x:x+w]  # Crop the face

        # Resize the face to a fixed size for consistency
        cropped_face_bgr = cv2.cvtColor(cropped_face, cv2.COLOR_RGB2BGR)
        cropped_face_resized = cv2.resize(cropped_face_bgr, (160, 160))  # Resize for FaceNet model

        # Display the cropped face to the user
        cv2.imshow("Cropped Face - Confirm Identity", cropped_face_resized)
        print("Cropped face displayed. Please confirm.")

        # Wait for user input to confirm or skip
        key = cv2.waitKey(0) & 0xFF
        if key == ord('q'):  # Press 'q' to quit without adding
            print("Operation cancelled by the user.")
            cv2.destroyWindow("Cropped Face - Confirm Identity")
            continue

        # Generate embedding for the face
        try:
            embedding = embedder.embeddings([cropped_face_resized])[0]
        except Exception as e:
            print(f"Error generating embedding: {e}")
            continue

        # Check if the face already exists in the database
        existing_name, distance = is_face_in_database(embedding)
        if existing_name:
            print(f"This face matches with {existing_name} (Distance: {distance:.2f}).")
            is_same = input(f"Is this {existing_name}? (yes/no): ").strip().lower()
            if is_same == 'yes':
                print(f"{existing_name} is already in the database. Skipping addition.")
                cv2.destroyWindow("Cropped Face - Confirm Identity")
                continue

        # Ask for a name for the face
        name = input("Enter the name for this face (or type 'skip' to skip): ").strip()
        cv2.destroyWindow("Cropped Face - Confirm Identity")

        if name.lower() == "skip":
            print("Face skipped.")
            continue

        # Add the new face to the database
        # Store the embedding and initialize an empty emotion dictionary
        known_faces[name] = (embedding, {})  # Store embedding and an empty emotion data dictionary
        print(f"Added {name} to the database.")

# Function to analyze emotion
def analyze_emotion(face_image):
    """Analyze the emotion of a given face image."""
    try:
        analysis = DeepFace.analyze(face_image, actions=['emotion'], enforce_detection=True)
        return analysis[0]['dominant_emotion']
    except Exception as e:
        print(f"Error in emotion analysis: {e}")
        return None

# Function to log attendance in the Excel file
def log_attendance(name, action):
    """Log attendance for the recognized face (entry or exit)."""
    sheet = attendance_workbook.active
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append([name, timestamp, action])
    attendance_workbook.save(ATTENDANCE_FILE)
    print(f"Attendance for {name} ({action}) logged at {timestamp}.")

# Function to capture a photo from the webcam and add it to the database
def capture_photo_and_add_to_database():
    """Capture a photo from the webcam and add it to the database."""
    cap = cv2.VideoCapture(0)

    if not cap.isOpened():
        print("Error: Could not open video device.")
        return

    print("Press 'c' to capture a photo or 'q' to quit.")
    
    while True:
        ret, frame = cap.read()
        if not ret:
            print("Failed to capture frame. Exiting...")
            break

        cv2.imshow("Webcam", frame)

        key = cv2.waitKey(1) & 0xFF
        if key == ord('c'):  # Capture photo
            rgb_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            detections = detector.detect_faces(rgb_image)

            if len(detections) == 0:
                print("No face detected. Please try again.")
                continue

            for detection in detections:
                x, y, w, h = detection['box']
                x, y = max(0, x), max(0, y)
                cropped_face = rgb_image[y:y+h, x:x+w]  # Crop the face

                cropped_face_bgr = cv2.cvtColor(cropped_face, cv2.COLOR_RGB2BGR)
                cropped_face_resized = cv2.resize(cropped_face_bgr, (160, 160))  # Resize for FaceNet model

                cv2.imshow("Cropped Face - Confirm Identity", cropped_face_resized)
                print("Cropped face displayed. Please confirm.")

                key = cv2.waitKey(0) & 0xFF
                if key == ord('q'):  # Quit without adding
                    print("Operation cancelled by the user.")
                    cv2.destroyAllWindows()
                    continue

                # Generate embedding for the face
                try:
                    embedding = embedder.embeddings([cropped_face_resized])[0]
                except Exception as e:
                    print(f"Error generating embedding: {e}")
                    continue

                # Check if the face already exists in the database
                existing_name, distance = is_face_in_database(embedding)
                if existing_name:
                    print(f"This face matches with {existing_name} (Distance: {distance:.2f}).")
                    is_same = input(f"Is this {existing_name}? (yes/no): ").strip().lower()
                    if is_same == 'yes':
                        print(f"{existing_name} is already in the database. Skipping addition.")
                        cv2.destroyAllWindows()
                        continue

                # Detect emotion
                emotion = detect_emotion(cropped_face_bgr)
                if emotion:
                    print(f"Detected emotion: {emotion}")
                else:
                    print("Emotion detection failed.")

                # Ask for a name for the face
                name = input("Enter the name for this face (or type 'skip' to skip): ").strip()
                cv2.destroyAllWindows()

                if name.lower() == "skip":
                    print("Face skipped.")
                    continue

                # Add the new face to the database
                known_faces[name] = (embedding, {"emotion": emotion})  # Store embedding and emotion
                print(f"Added {name} to the database with emotion: {emotion}")
                save_database()  # Save the updated database

        elif key == ord('q'):  # Quit the webcam capture
            print("Exiting the webcam capture.")
            break

    cap.release()
    cv2.destroyAllWindows()

# Function to recognize faces and log attendance
def recognize_faces_and_log_attendance():
    """Recognize faces and log attendance in real-time using the webcam."""
    cap = cv2.VideoCapture(0)

    if not cap.isOpened():
        print("Error: Could not open video device.")
        return

    entry_logged = {}
    exit_logged = set()  # Track people who have logged "Exit"

    while True:
        ret, frame = cap.read()
        if not ret:
            print("Failed to capture frame. Exiting...")
            break

        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        detections = detector.detect_faces(rgb_frame)

        for detection in detections:
            x, y, w, h = detection['box']
            x, y = max(0, x), max(0, y)
            cropped_face = rgb_frame[y:y+h, x:x+w]

            # Generate embedding for the detected face
            try:
                embedding = embedder.embeddings([cv2.resize(cropped_face, (160, 160))])[0]
                name, distance = is_face_in_database(embedding)

                if name:
                    emotion = analyze_emotion(cropped_face)
                    cv2.putText(frame, f"{name} - {emotion}", (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)

                    # Log entry only the first time
                    if name not in entry_logged:
                        log_attendance(name, "Entry")
                        entry_logged[name] = True

                    # Log exit only once
                    if name not in exit_logged:
                        cv2.putText(frame, "Press 'e' to exit", (x, y + h + 10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 0, 0), 2)

                        # Detect exit event (press 'e' key)
                        if cv2.waitKey(1) & 0xFF == ord('e'):  # Press 'e' to log exit
                            log_attendance(name, "Exit")
                            exit_logged.add(name)

                else:
                    cv2.putText(frame, "Unknown", (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)

            except Exception as e:
                print(f"Error processing face: {e}")

        cv2.imshow("Real-Time Face Recognition", frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

# Main function to run the application
def main():
    """Main function to run the face recognition application."""
    while True:
        print("\nOptions:")
        print("1. Add face from file")
        print("2. Recognize faces and log attendance")
        print("3. Save database")
        print("4. ADD From webcam")
        print("5. Exit")
        choice = input("Choose an option: ").strip()

        if choice == '1':
            file_path = filedialog.askopenfilename(title="Select an Image", filetypes=[("Image Files", "*.jpg;*.jpeg;*.png")])
            if file_path:
                add_to_database_from_file(file_path)
        elif choice == '2':
            recognize_faces_and_log_attendance()
        elif choice == '3':
            save_database()
        elif choice == '4':
            capture_photo_and_add_to_database()    
        elif choice == '5':
            save_database()
            print("Exiting the application.")
            break
        else:
            print("Invalid choice. Please try again.")

if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()  # Hide the root window
    main()
